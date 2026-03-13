from openpyxl import load_workbook

def get_login_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active

    login_list = []
    common_password = "testingisfun99"

    for row in sheet.iter_rows(min_row=2, values_only=True):
        username = row[0]

        if username in ["demouser", "image_not_loading_user", "fav_user"]:
            login_list.append((username, common_password))

    return login_list